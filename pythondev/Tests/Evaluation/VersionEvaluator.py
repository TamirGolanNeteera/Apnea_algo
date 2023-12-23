import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))  # noqa

from Tests.Utils.PathUtils import generate_plot_path, get_folder_list, \
    find_benchmarks_from_folders, remove_sep_at_end
from Tests.Utils.StringUtils import string_match_out_of_list, subtract_strings
from pylibneteera.math_utils import normal_round

from Configurations import spot_config

from Tests.Utils.PandasUtils import nan_first, add_row_to_df, add_tilda, get_common_rows, \
    find_first_entry_of_multi_index, is_time_series
from Tests.Utils.NumpyUtils import weighted_average
from Tests.Utils.DBUtils import get_cam_hyperlink, match_lists_by_ts, setup_types, get_not_invalid_setups, get_bmi
from Tests.Evaluation.EvaluationUtils import under_thresh, under_percent, under_percent_or_thresh
from Tests.Utils.LoadingAPI import load_pred_rel_from_npy, load_pred_high_qual, get_list_of_setups_in_folder_from_vs, \
    get_setups_union_folders, get_setups_intersect_folders, ReferenceHolder
from Tests.Evaluation.PPTutils import create_compare_fig_ppt, create_compare_ppt_between_radars
from Tests.Plots.PlotResults import main_plot_results, plot_bland_altman
from Tests.Utils.ResearchUtils import print_var  # noqa
from Tests.Utils.TestsUtils import *
from Tests.Constants import *
import Tests.bbi.BBIPerformance as BBIP

from sklearn.metrics import r2_score, confusion_matrix
from argparse import RawTextHelpFormatter
import argparse
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
from shutil import copy2
import warnings
import xlsxwriter  # noqa
import copy
import numpy as np
import os
import traceback
import matplotlib
from itertools import cycle

warnings.filterwarnings("ignore")


def get_args() -> argparse.Namespace:
    """ Parse arguments"""
    parser = argparse.ArgumentParser(description='Result Evaluator Post Process', formatter_class=RawTextHelpFormatter)
    parser.add_argument('-folder_list', '-folders_list', metavar='folder_list', type=str, nargs='+', required=True,
                        help='list of csv files to be evaluated')
    parser.add_argument('-setups', '-session_ids', metavar='ids', nargs='+', type=int,
                        help='Setup IDs in DB to collect the online results', required=False)
    parser.add_argument('-version_names', '-versions', type=str, nargs='+', required=False,
                        help='list of algo version names')
    parser.add_argument('-vital_signs', '-compute', '-vital_sign', '-vs', metavar='vital_sign', type=str, nargs='+',
                        choices=['hr', 'rr', 'bbi', 'stat', 'ie'], default=['hr', 'rr', 'stat', 'bbi', 'ie'],
                        help='compute the following vital signs')
    parser.add_argument('-fname', '-file', type=str, required=False, help='name of output file')
    parser.add_argument('-start_from_second', metavar='start_from_second', type=int, required=False, default=60,
                        help='the second that the real measurement starts at')
    parser.add_argument('-output_path', metavar='output_path', type=str, required=False,
                        help='directory to save the output files in')
    parser.add_argument('-match_list_type', metavar='match_lists', type=str, default='ts',
                        help='what type of match list to use', choices=['NO', 'MSE', 'TS', 'no', 'ts', 'mae'])
    parser.add_argument('--t0', action='store_true', help='start each setup from its t0')
    parser.add_argument('-transition_length', type=int, default=20, help='ignore transition area for stat evaluation')
    parser.add_argument('--force', action='store_true', help='ignore validity of reference and radar data')
    parser.add_argument('--force_ref', action='store_true', help='ignore validity of reference')
    parser.add_argument('-outputs', '-output', type=str, nargs='+', default=['pptx', 'xlsx', 'bland_altman'],
                        choices=['pptx', 'xlsx', 'bland_altman'], help='list of comparison output file types')
    parser.add_argument('--include_negative', action='store_true', help='use all time points including the negative')
    parser.add_argument('--plot', '--plot_ref', action='store_true', help='run PlotResults on setups')
    parser.add_argument('--best_setup_for_session', action='store_true',
                        help='also generate excel with 1 best setup per session')
    parser.add_argument('--merge_benchmarks', action='store_true', help='merge multiple folders statistics')
    parser.add_argument('--dont_overwrite', action='store_true', help='keep the saved exel if exists')
    parser.add_argument('-features', type=str, nargs='+', default=[], help='aggregate by feature values')

    # for PlotResults.py :
    parser.add_argument('--diff', action='store_true', help='Plot the pred-ref differences')
    parser.add_argument('-ppt_fname', type=str, required=False, help='Powerpoint filename to generate ppt '
                                                                     'presentation from plots')
    parser.add_argument('-product', type=str, required=False, default='health', help='automotive or health')

    return parser.parse_args()


def set_sheet_visualization(freeze_pane_cell, title_row, sheet_name, writer, xl_workbook, summary=False):
    sheet = writer.sheets[sheet_name]  # pull sheet object
    xl_workbook.active = xl_workbook.sheetnames.index('Welcome')
    sheet.freeze_panes = sheet[freeze_pane_cell]

    rows_generator = sheet.iter_rows(min_row=title_row)
    for i, cell in enumerate(next(rows_generator)):  # iterate over 2nd row
        if cell.value is not None:
            sheet.column_dimensions[cell.column_letter].width = max([len(y) for y in cell.value.split()] + [7]) + 3
            cell.alignment = Alignment(wrap_text=True)
            if cell.value in ['Video', 'Plot', 'path']:
                for video_cell in next(sheet.iter_cols(min_col=i + 1, min_row=3)):
                    video_cell.style = 'Hyperlink'
    if summary:
        sheet['B1'].value = sheet_name.replace('_', ' ')
        sheet['B1'].font = Font(bold=True, sz=14)
        if freeze_pane_cell == 'C3':
            for cell in next(rows_generator):
                cell.alignment = Alignment(wrap_text=True)
        for cell in next(sheet.iter_cols()):
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 26
        sheet.column_dimensions['B'].width = 18
        for col in sheet.iter_cols(min_col=3):
            if col[0].value is None or np.any(
                    [x.lower() in col[0].value.lower() for x in
                     ['fpr', 'true', 'under', 'out of', 'percent', 'rate', 'percentage']]):
                cell_format = '0.00%'
            elif 'num' in col[0].value.lower():
                cell_format = '0'
            else:
                cell_format = '0.0'
            for cell in col[1:]:
                cell.number_format = cell_format
    else:
        sheet.auto_filter.ref = sheet.dimensions.replace('A1', 'A2')
        for col in sheet.iter_cols(min_col=2):
            if col[1].value and ('true positive rate tolerance' in col[1].value or 'perc' in col[1].value.lower()):
                for cell in col[2:]:
                    cell.number_format = '0.0%'
            elif isinstance(col[2].value, float) and col[2].value % 1 > 1e-4:
                for cell in col[2:]:
                    cell.number_format = '0.0'


def get_db_columns(setups: list, db) -> pd.DataFrame:
    """
    get the dbview (as the php page) data and the setup type of the setup

    """
    if len(setups) == 0:
        return pd.DataFrame()
    df = pd.DataFrame(index=np.arange(len(setups)) + 1)
    setup_type_column = []
    all_types = setup_types(db)
    db_view = [db.setup_view(setup)[0] for setup in setups]
    for setup in setups:
        type_found = False
        for setup_type in all_types.keys():
            if setup in all_types[setup_type]:
                setup_type_column.append(setup_type)
                type_found = True
                break
        if not type_found:
            setup_type_column.append('')
    df['Setup Type'] = setup_type_column
    df['Video'] = [get_cam_hyperlink(setup, db) for setup in setups]
    df.index = np.arange(len(df))
    df = pd.concat([df, pd.DataFrame(db_view)], axis=1)
    df['time'] = df['time'].dt.strftime('%d/%m/%y %H:%M')
    df['path'] = [f'=HYPERLINK(\"{x}\")' for x in df['path']]
    df['BMI'] = [get_bmi(setup, db) for setup in setups]
    return df.set_index(df.setup.astype(int))


def open_excel_and_get_writer(path, dont_overwrite=True):
    # change_dir_to_pythondev()
    in_progress_path = path.replace('.xlsx', '_in_progress.xlsx')
    if dont_overwrite and os.path.exists(path):
        copy2(path, in_progress_path)
    else:
        copy2(os.path.join(os.path.dirname(__file__), 'Welcome_page_version_eval.xlsx'), in_progress_path)
    exel_workbook = load_workbook(in_progress_path)
    writer_pandas = pd.ExcelWriter(in_progress_path, engine='openpyxl')
    writer_pandas.book = exel_workbook
    return writer_pandas, exel_workbook


def save_df_to_excel(df, ver, write, **kwargs):
    """save the dataframe to excel with the given name on top"""
    name_ver = pd.DataFrame({'v': [ver]})
    name_ver.to_excel(write, startrow=0, header=None, index=None, **kwargs)
    df.to_excel(write, startrow=1, index=None, **kwargs)


def create_full_setups_sheet(dfs, name, db_view):
    current_column = 0
    for i, (ver_name, df) in enumerate(dfs.items()):
        temp_df = df.sort_values('Setup #')
        if i == 0:
            if len(dfs) > 1 and 'Absolute Mean Error [bpm]' in temp_df:
                temp_df['error diff'] = \
                    temp_df['Absolute Mean Error [bpm]'] - list(dfs.values())[1]['Absolute Mean Error [bpm]']
        else:
            temp_df.drop(['Setup #'], axis=1, inplace=True)
        save_df_to_excel(temp_df, ver_name, writer, sheet_name=name, startcol=current_column)
        current_column += temp_df.shape[1] + 1
    save_df_to_excel(db_view, 'db_view', writer, sheet_name=name, startcol=current_column)
    set_sheet_visualization('B3', 2, name, writer, xl_workbook, False)


def compare_between_distances(data_frame, db_view):
    df = pd.concat((data_frame, db_view), axis=1)
    distance_data_frames = {f'distance = {dist}mm': df[df.distance == dist].set_index('session')
                            for dist in sorted(df.distance.unique())}
    distance_data_frames = get_common_rows(distance_data_frames)
    return distance_data_frames


def args_to_only_high_qual(arg_object):
    args_with_high_qual = copy.deepcopy(arg_object)
    args_with_high_qual.only_high_qual = True
    return args_with_high_qual


def args_to_only_low_qual(arg_object):
    new_args = copy.deepcopy(arg_object)
    new_args.only_low_qual = True
    new_args.only_high_qual = False
    return new_args


def get_full_file_path(output_path, folders, fname, is_t0, start_from, match_type, version_names, merge_benchmarks,
                       force):
    """create the output file name"""
    out_folder = output_path
    first_folder = folders[0]
    if out_folder is None:
        out_folder = first_folder
        if out_folder.endswith(os.sep):
            out_folder = out_folder[:-1]
        if out_folder.split(os.sep)[-1] in ['spot', 'dynamic']:
            out_folder = os.path.dirname(out_folder)
        if 'optimization' in first_folder:
            out_folder = os.path.dirname(out_folder)
    if merge_benchmarks:
        out_folder = os.path.dirname(out_folder)
    if fname is None:
        if merge_benchmarks:
            file_name = f'eval_{version_names[0]}' + find_benchmarks_from_folders(folders)
        else:
            file_name = ('eval_' if len(version_names) == 1 else 'compare_') \
                        + join(version_names, '_') + find_benchmarks_from_folders(folders)
    else:
        file_name = fname
    file_name = file_name.replace('_selected_sessions', '')
    start_from_str = '_using_t0' if is_t0 else f'_starts_from_sec_{start_from}'
    if force:
        start_from_str += '_using_force'
    xl_file_name = file_name + f'{start_from_str}_{match_type}_match_lists.xlsx'
    return os.path.join(out_folder, xl_file_name), out_folder, file_name


def validate_args(args):
    args.match_list_type = args.match_list_type.lower().replace('mse', 'mae')
    if args.folder_list is None:
        args.folder_list = [args.output_path]
    args.folder_list, out_path = get_folder_list(args.folder_list)
    if args.output_path is None:
        args.output_path = out_path
    if 'nwh' in str(args.folder_list[0]):
        args.force_ref = True
    if 'ec_benchmark' in str(args.folder_list[0]):
        args.vital_signs = ['stat']
    if args.version_names is None:
        args.version_names = [get_ver_name_from_path(folder) for folder in args.folder_list]

    args.full_file_path, args.output_path, args.fname = get_full_file_path(
        args.output_path, args.folder_list, args.fname, args.t0,
        args.start_from_second, args.match_list_type, args.version_names, args.merge_benchmarks, args.force)

    if np.all([x.isnumeric() for x in args.version_names]):
        args.version_names = sorted(np.array(args.version_names, dtype=float))
    args.only_high_qual = False
    args.only_low_qual = False
    return args


def get_high_qual_no_high_qual_spot(mut_args):
    output = []
    list_dir = os.listdir(mut_args.folder_list[0])
    dir_split = mut_args.folder_list[0].split(os.path.sep)
    if ('spot' in list_dir or 'dynamic' in list_dir) and 'spot' not in dir_split and 'dynamic' not in dir_split:
        for mode in intersect([['spot', 'dynamic'], list_dir]):
            arg_copy = copy.deepcopy(mut_args)
            arg_copy.folder_list = [os.path.join(folder, mode) for folder in arg_copy.folder_list]
            output.append(arg_copy)
            if mode == 'dynamic':
                output.append(args_to_only_high_qual(arg_copy))
                output.append(args_to_only_low_qual(mut_args))
    else:
        output.append(mut_args)
        output.append(args_to_only_high_qual(mut_args))
        output.append(args_to_only_low_qual(mut_args))
    return output


def get_list_of_args(orig_args):
    mutual_args = validate_args(orig_args)
    list_of_arg_objects = get_high_qual_no_high_qual_spot(mutual_args)
    return list_of_arg_objects


def get_ver_name_from_path(path):
    set_of_benchmarks = {str(x) for x in Benchmark} | {str(x) for x in Project}
    if os.path.basename(path) == '':
        path = os.path.dirname(path)
    if os.path.basename(path) in ['dynamic', 'spot']:
        path = os.path.dirname(path)
    if os.path.basename(path) in set_of_benchmarks:
        path = os.path.dirname(path).replace('szmc_clinical_trials', 'szmc')
    if os.path.basename(path) == 'py':
        path = os.path.dirname(path)
    if os.path.basename(path) == 'stats':
        path = os.path.dirname(path)
    benchmark = string_match_out_of_list(os.path.basename(path), set_of_benchmarks)
    if benchmark:
        return subtract_strings(os.path.basename(path), benchmark)

    return os.path.basename(path)


def database_by_session(df):
    if df is None:
        return None
    output = df.groupby('session').aggregate(list)
    output['session'] = output.index
    return output


class EvaluatorWrapper:
    def __init__(self, db_object, ref_holder: ReferenceHolder):
        self.database_df = None
        self.sheet_name = None
        self.best_setups = {}
        self.db = db_object
        self.ref_holder = ref_holder

    def eval_single_vs(self, args_list, vs):
        evaluator_vs = {'hr': HREvaluator,
                        'rr': RREvaluator,
                        'ie': IEEvaluator,
                        'stat': STATEvaluator,
                        'bbi': BBIEvaluator}[vs](self, args_list, vs)
        if evaluator_vs.plot:
            evaluator_vs.run_plotter()
        evaluator_vs.evaluate()
        self.best_setups = evaluator_vs.best_setups


class HREvaluator(EvaluatorWrapper):
    def __init__(self, wrapper, l_args, v):
        super().__init__(None, None)
        self.summary_hiqh_qual_low_qual_spot_all_df = None
        self.summary_first_metric_per_radar_spot_and_dynamic = None
        self.sheet_name = None
        self.vs = v
        self.args_list = l_args
        self.update_args(wrapper.__dict__)
        self.low_value = VERY_LOW_VALUES.get(self.vs)
        self.high_value = VERY_HIGH_VALUES.get(self.vs)
        self.units = UNITS.get(self.vs)
        self.number_of_plots_comparison = 0
        self.relevant_setups = None
        self.update_args(l_args[0].__dict__)
        self.pred = {}
        self.high_qual = {}
        self.spot = {}
        self.load_ref()
        self.load_pred()
        self.metric_list = UNDER_DICT.get(self.vs)
        first_folder = self.folder_list[0]
        self.is_spot = 'spot' in [os.path.basename(first_folder), os.path.dirname(first_folder)]
        self.data_frames = None
        self.limits = VITAL_SIGN_LIMITS.get(v)
        self.all_pred = dict()
        self.all_gt = dict()

    def update_args(self, args_dict):
        for name in args_dict:
            setattr(self, name, args_dict[name])

    def get_df_by_category(self, df, category):
        """
        gets the rows from df
        """
        if 'Setup Type' not in df:
            df = pd.concat([df, self.database_df], axis=1)
            if 'number of time points' in df:
                df = df[~df['number of time points'].isna()]
        if self.low_value is not None and category == f'~Under {self.low_value} bpm':
            df = df.loc[df['Median Ground Truth'] <= self.low_value]
            df = df.loc[~df['Setup #'].isin(db.setup_by_posture(Posture.standing))]
        elif self.low_value is not None and f'Above {self.high_value} bpm' in category:
            df = df.loc[df['Median Ground Truth'] >= self.high_value]
            df = df.loc[~df['Setup #'].isin(db.setup_by_posture(Posture.standing))]
            if 'Back' in category:
                df = df.loc[df['Setup #'].isin(db.setup_by_target(Target.back))]
            else:
                df = df.loc[df['Setup #'].isin(db.setup_by_target(Target.front))]
        elif category.replace('~', '') in ULTRA_CATEGORIES:
            df = df.loc[df['Setup Type'].str.contains(category.replace('~', ''))]
        elif category != 'All':
            df = df.loc[df['Setup Type'] == category]
        return df

    def get_existing_categories(self, categories: list):
        """
        getting the categories from list and the ultra categories and all
        """
        unique = list(set(categories))
        all_categories = ['All']
        if self.low_value is not None:
            all_categories += [f'~Under {self.low_value} bpm', f'~Back Above {self.high_value} bpm',
                               f'~Front Above {self.high_value} bpm']
        for category in unique:
            if category == '':
                all_categories.append('~~Not found')
            elif category not in ULTRA_CATEGORIES:
                all_categories.append(category)
            for ult_cat in ULTRA_CATEGORIES:
                if ult_cat in category and ult_cat not in all_categories:
                    all_categories.append('~' + ult_cat)
        return all_categories

    def run_plotter(self, setups=None):
        args_for_plot = argparse.Namespace(**self.__dict__)
        args_for_plot.setups = None if setups is None else list(setups.keys())
        args_for_plot.vital_sign = [self.vs]
        for folder in self.folder_list:
            if len(re.findall('net-alg-[0-9]+', folder)) == 0:
                args_for_plot.save_fig_path = None
                args_for_plot.result_dir = folder
                main_plot_results(args_for_plot, db)

    def plot_and_compare_selected_setups(self, dfs):
        selected_setups = {}
        min_diffs = {'Absolute Mean Error [bpm]': 1, 'Median Prediction': 5, 'positive time points': 15}
        time_under_min_diff = cycle(UNDER_DICT[self.vs])
        for i, col in enumerate(list(dfs.values())[0].columns):
            if 'Time points under' in col:
                metric_df = pd.concat(
                    [df[col] / (df['positive time points'] + 1e-8) for df in dfs.values()], axis=1)
                diffs = metric_df.max(axis=1) - metric_df.min(axis=1)
                min_diff = next(time_under_min_diff).get('diff')
                if min_diff is not None:
                    diffs = diffs[diffs > min_diff]
                else:
                    continue
            elif col in min_diffs:
                metric_df = pd.concat([df[col] for df in dfs.values()], axis=1)
                diffs = metric_df.max(axis=1) - metric_df.min(axis=1)
                diffs = diffs[diffs > min_diffs[col]]
            else:
                continue
            for setup in list(diffs.abs().sort_values(ascending=False).index[:10]):
                diff_str = f'{diffs[setup] * 100:.1f}%' if 'Time points under' in col else f'{diffs[setup]:.3f}'
                selected_setups[setup] = selected_setups.get(setup, '') + f'{col}: diff={diff_str}\n'
        if len(selected_setups):
            self.run_plotter(selected_setups)
            out_path = os.path.join(self.output_path, f'{self.vs}_greatest_change_{self.fname}.pptx')
            create_compare_fig_ppt(self.folder_list, self.version_names, out_path,
                                   self.vs, selected_setups.keys(), self.database_df, metadata=selected_setups)
        self.number_of_plots_comparison = len(selected_setups)

    def throw_pred_gt_out_of_limits(self, pred_gt_high_qual_df: pd.DataFrame):
        pred_gt_df = pred_gt_high_qual_df.loc[:, ['pred', 'gt']]

        if self.include_negative:
            self.limits['min'] = -2
        valid_indices = np.logical_and(self.limits['min'] < pred_gt_df, pred_gt_df < self.limits['max']).all(axis=1)
        return pred_gt_high_qual_df[valid_indices]

    def match_lists(self, v_name, setup):
        pred = self.pred[v_name][setup]
        high_qual = self.high_qual[v_name][setup]
        gt = self.ref_holder.get_reference(self.vs, setup)
        if high_qual is None:
            high_qual = np.zeros(len(pred))

        pred_high_qual = pd.DataFrame({'pred': pred, 'high_qual': high_qual})
        if self.match_list_type == 'no' or is_time_series(gt):
            min_len = min(len(gt), len(pred))
            pred_high_qual, gt = pred_high_qual[:min_len], gt[:min_len]
            shift = 0
        elif self.is_spot or self.match_list_type == 'ts':
            pred_high_qual, gt, shift = match_lists_by_ts(pred_high_qual, gt, setup_id=setup, vs=self.vs, db=db)
            min_len = min(len(gt), len(pred_high_qual))
            if len(pred_high_qual) > len(gt):
                print(f'setup {setup} vs {self.vs} pred finished after gt')
            pred_high_qual, gt = pred_high_qual[:min_len], gt[:min_len]
        elif self.match_list_type == 'mse':
            pred_high_qual, gt, shift = match_lists(pred_high_qual, gt)
        else:
            raise ValueError
        if len(pred) == 0:
            raise ValueError(f'matched pred and ref is empty. setup {setup}, vital sign: {self.vs}')
        pred_high_qual['gt'] = gt
        if self.is_spot:
            pred_high_qual = pd.DataFrame(pred_high_qual.iloc[-60:, :].median(axis=0).apply(normal_round)).T
            pred_high_qual['pred'] = load_pred_high_qual(
                self.folder_list[self.version_names.index(v_name)], setup, self.vs)['pred']
        return pred_high_qual, shift

    def get_start_time(self, setup_id, shift):
        start_from_second = self.start_from_second
        if self.t0:
            try:
                start_from_second = db.setup_spot(setup_id)['t0']
            except KeyError:
                pass
        if shift < 0:
            start_from_second += shift
        return max(start_from_second, 0)

    def match_pred_high_qual_gt(self, idx, name):
        pred_gt_df, delay = self.match_lists(name, idx)
        if not self.is_spot:
            start_from = self.get_start_time(idx, delay)
            pred_gt_df = pred_gt_df.iloc[start_from:, :]
        pred_gt_df['diff'] = np.abs(pred_gt_df['pred'] - pred_gt_df['gt'])
        return pred_gt_df, delay

    def get_time_points_under(self, metric, df):
        percent = metric.get('per')
        thresh = metric.get('thresh')
        assert percent or thresh
        if percent is None:
            return f'Time points under {thresh}{self.units} error', np.sum(under_thresh(df['diff'], thresh))
        if thresh is None:
            return f'Time points under {percent}% error', np.sum(under_percent(df['diff'], df['gt'], percent))
        return f'Time points under {percent}% error or {thresh}{self.units} error', \
               np.sum(under_percent_or_thresh(df['diff'], df['gt'], thresh, percent))

    def evaluate_single_setup(self, setup, ver):
        pred_gt_high_qual, shift_sec = self.match_pred_high_qual_gt(setup, ver)
        setup_series = pd.Series({'Setup #': setup,     # todo
                                  'number of time points': len(pred_gt_high_qual),
                                  'positive time points': np.sum(pred_gt_high_qual['pred'] > 0),
                                  })
        gt = pred_gt_high_qual['gt']
        setup_series['Median Ground Truth'] = gt.median()
        setup_series['Min Ground Truth'] = gt.min()
        setup_series['Max Ground Truth'] = gt.max()
        pred_gt_high_qual = self.throw_pred_gt_out_of_limits(pred_gt_high_qual)
        if not self.is_spot:
            if self.only_high_qual:
                setup_series['time until high quality'] = first_ind(self.high_qual[ver][setup])
                setup_series['high_quality time points'] = pred_gt_high_qual['high_qual'].sum()
                pred_gt_high_qual = pred_gt_high_qual[pred_gt_high_qual['high_qual'] == 1]
            elif self.only_low_qual:
                setup_series['low quality time points'] = \
                    len(pred_gt_high_qual['high_qual']) - pred_gt_high_qual['high_qual'].sum()
                pred_gt_high_qual = pred_gt_high_qual[pred_gt_high_qual['high_qual'] == 0]
            elif not self.include_negative:
                pred_gt_high_qual = pred_gt_high_qual[pred_gt_high_qual['pred'] >= 0]
            setup_series[f'Error Standard deviation [{self.units}]'] = pred_gt_high_qual['diff'].std()
        if self.is_spot or self.only_high_qual:
            self.all_pred[ver] = self.all_pred.get(ver, []) + list(pred_gt_high_qual['pred'])
            self.all_gt[ver] = self.all_gt.get(ver, []) + list(pred_gt_high_qual['gt'])

        setup_series['time points in limits'] = len(pred_gt_high_qual)
        for i, under_metric in enumerate(self.metric_list):
            key, val = self.get_time_points_under(under_metric, pred_gt_high_qual)
            setup_series[key] = val
            if i == 0 and not self.is_spot:
                setup_series[f'Percentage {key}'] = val / setup_series['time points in limits']
        setup_series['Median Prediction'] = pred_gt_high_qual['pred'].median()
        setup_series['First Prediction'] = nan_first(pred_gt_high_qual['pred'])
        setup_series['First Ground Truth'] = nan_first(gt)
        setup_series[f'Absolute Mean Error [{self.units}]'] = pred_gt_high_qual['diff'].mean()
        setup_series['shift'] = shift_sec

        return setup_series

    def evaluate_single_path_files(self, ver_name, folder):
        df = None
        for setup_num in self.relevant_setups:
            setup_results = self.evaluate_single_setup(setup_num, ver_name)
            if setup_results is not None:
                df = add_row_to_df(df, setup_num, setup_results)
            else:
                print(f'setup {setup_num} vs {self.vs} skipped')
        if df is None or len(df) == 0:
            raise FileNotFoundError()

        df['Plot'] = [generate_plot_path(folder, self.vs, setup) for setup in df.index]
        return df

    def generate_evaluation_sheet(self) -> (Optional[pd.DataFrame], Optional[pd.DataFrame]):
        print(f'evaluating {self.sheet_name}')
        data_frames = dict()
        try:
            for version_name, path in zip(self.version_names, self.folder_list):
                data_frames[version_name] = self.evaluate_single_path_files(version_name, path)
        except FileNotFoundError:
            print(f' {self.sheet_name} not found, skipping to next vs')
            return None

        data_frames = get_common_rows(data_frames)
        setup_list = list(data_frames.values())[0].index.to_list()
        relevant_df_db = self.ger_relevant_db_view(setup_list)
        create_full_setups_sheet(data_frames, self.sheet_name, relevant_df_db)
        return data_frames

    def ger_relevant_db_view(self, setups):
        if self.database_df is None:
            df = get_db_columns(setups, db).sort_index()
            self.database_df = df
        else:
            setups_in_df_db = intersect([setups, self.database_df.index.to_list()])
            setups_not_in_df_db = list(set(setups) - set(setups_in_df_db))
            df = pd.concat([self.database_df.loc[setups_in_df_db], get_db_columns(setups_not_in_df_db, db)])
            df.sort_index(inplace=True)
        return df

    def generate_sheet_name(self):
        sheet_name = self.vs  # + f'_start_{self.start_from_second}'
        if 'spot' in self.folder_list[0]:
            sheet_name += '_spot'
        else:
            if self.only_high_qual:
                sheet_name += '_only_high_qual'
            elif self.only_low_qual:
                sheet_name += '_only_low_qual'
        self.sheet_name = sheet_name

    def summarize_version_category_dynamic(self, df):
        index = pd.MultiIndex(levels=[[], []], codes=[[], []])
        summary = pd.Series(index=index)
        summary['number of points for statistics'] = df['time points in limits'].sum()
        summary['positive time points percentage'] = \
            df['positive time points'].sum() / df['number of time points'].sum()
        if self.only_high_qual:
            summary['mean time for high quality'] = df['time until high quality'].mean()
            summary['high_quality time points out of non-negative points'] = \
                df['high_quality time points'].sum() / df['positive time points'].sum()
        df_clean = df[df['positive time points'] > 0]

        for metric in df_clean:
            if 'Time points under' in metric and 'Perc' not in metric:
                summary_key = metric.replace('Time points', '')
                try:
                    summary[(summary_key, 'per time point')] = \
                        df_clean[metric].sum() / df_clean['time points in limits'].sum()
                    under_time_pont_percent = 100 * df_clean[metric] / df_clean['time points in limits']
                    for percent_per_setup_thresh in PER_SETUP_THRESHOLDS:
                        over_thresh = under_time_pont_percent >= percent_per_setup_thresh
                        over_thresh[under_time_pont_percent.isna()] = np.nan
                        summary[(summary_key, f'{percent_per_setup_thresh}% per setup')] = \
                            np.sum(over_thresh) / over_thresh.notna().sum()
                except ZeroDivisionError:
                    summary[(summary_key, 'per time point')] = np.nan
                    for percent_per_setup_thresh in PER_SETUP_THRESHOLDS:
                        summary[(summary_key, f'{percent_per_setup_thresh}% per setup')] = np.nan

        summary['Error Standard deviation Average'] = \
            weighted_average(df_clean[f'Error Standard deviation [{self.units}]'], df_clean['time points in limits'])
        return summary

    @staticmethod
    def summarize_version_category_spot(df):
        summary = pd.Series()
        summary['Percentage of positive outputs'] = float(np.mean(df['Median Prediction'] > -1))
        df_clean = df.loc[:, :'Plot'].dropna()
        df_clean = df_clean[df_clean['Median Prediction'] != -1]
        for metric in df_clean:
            if 'Time points under' in metric:
                summary[metric.replace('Time points', '')] = df_clean[metric].mean()
        return summary

    def summarize_version_category(self, df):
        if self.is_spot:
            version_summary = self.summarize_version_category_spot(df)
        else:
            version_summary = self.summarize_version_category_dynamic(df)
        df_no_nans = df.dropna()
        version_summary['Absolute Mean Error Average'] = weighted_average(
            df_no_nans[f'Absolute Mean Error [{self.units}]'].fillna(0), df_no_nans['time points in limits'])
        if len(df_no_nans):
            version_summary['R-squared'] = r2_score(df_no_nans['Median Ground Truth'], df_no_nans['Median Prediction'])
        else:
            version_summary['R-squared'] = np.nan
        return version_summary

    @staticmethod
    def add_entry_at_start_of_series(series, key, value):
        if isinstance(series.index, pd.core.indexes.multi.MultiIndex):
            pre_metrics = pd.Series(value, index=pd.MultiIndex.from_tuples([(key, '')]))
        else:
            pre_metrics = pd.Series({key: value})
        return pd.concat([pre_metrics, series])

    @staticmethod
    def save_sum_to_excel(df, sheet, sheet_index=None):
        df.sort_index(level=0, sort_remaining=False, inplace=True)
        df.to_excel(writer, sheet_name=sheet, startrow=0)
        freeze_on_cell = 'C3' if isinstance(df.columns, pd.core.indexes.multi.MultiIndex) else 'C2'
        set_sheet_visualization(freeze_on_cell, 1, sheet, writer, xl_workbook, summary=True)
        if sheet_index is not None:
            sheets = xl_workbook._sheets
            sheet = sheets.pop()
            sheets.insert(sheet_index, sheet)

    def generate_summary(self, dframes: dict, sheet_name: str) -> pd.DataFrame:
        """
        creates data frame with row for each category over each version, and column for values
        """
        categories = self.get_existing_categories(self.database_df['Setup Type'])
        summary_df = None
        for category in categories:
            for version_name, df_raw in dframes.items():
                df = self.get_df_by_category(df_raw, category)
                if len(df) == 0:
                    continue
                version_summary = self.summarize_version_category(df)
                version_summary['Different Subjects number'] = len(np.unique(df['subject'].to_list()))
                if 'number of time points' in df:
                    version_summary['Average time point per setup'] = int(df['number of time points'].mean())
                version_summary = self.add_entry_at_start_of_series(version_summary, 'Num. Setups', df.shape[0])
                summary_df = add_row_to_df(summary_df, (category, version_name), version_summary)
                if category == 'All' and 'dist' not in sheet_name:
                    self.summary_hiqh_qual_low_qual_spot_all_df = add_row_to_df(
                        self.summary_hiqh_qual_low_qual_spot_all_df, (sheet_name, version_name), version_summary)
        self.save_sum_to_excel(summary_df, sheet_name)
        return summary_df

    def run_excels(self):
        data_frames = self.generate_evaluation_sheet()
        if data_frames is not None and len(list(data_frames.values())[0]) > 0:
            summary = self.generate_summary(data_frames, f'{self.sheet_name}_sum')
            self.data_frames = data_frames

    def run_presentations(self):
        out_path = os.path.join(self.output_path, f'{self.fname}_{self.vs}.pptx')

        if self.vs == 'hr' or 'hr' not in self.vital_signs:
            create_compare_ppt_between_radars(
                self.folder_list[0], 'distance', self.fname, self.vs, database_by_session(self.database_df))
            if len(self.folder_list) == 1:
                return
        create_compare_fig_ppt(self.folder_list, self.version_names, out_path, self.vs, self.setups,
                               db_view=self.database_df, min_length=self.number_of_plots_comparison + 1)

    def get_best_setup_each_session(self):
        best_dfs = dict()
        if len(best_dfs) == 0:
            for key, df in self.data_frames.items():
                df_w_db = pd.concat((list(self.data_frames.values())[0], self.database_df), axis=1)
                if len(self.best_setups) == 0:
                    best_criteria = [x for x in df.columns if 'Percentage' in x][0]
                    self.best_setups = df_w_db.groupby(
                        'session', sort=False)[best_criteria].transform(max) == df_w_db[best_criteria]
                best_dfs[key] = df_w_db[self.best_setups]

        sheet = writer.sheets[self.sheet_name]
        rows_generator = sheet.iter_rows(min_row=3)
        for row in rows_generator:
            if self.best_setups[int(row[0].value)]:
                for cell in row:
                    cell.font = Font(bold=True)
        return best_dfs

    def run_single_vs_single_settings(self):
        self.generate_sheet_name()
        self.run_excels()
        if not self.only_high_qual and not self.only_low_qual and self.data_frames is not None and not self.is_spot:
            df = list(self.data_frames.values())[0]
            self.plot_and_compare_selected_setups(self.data_frames)
            distance_dfs = compare_between_distances(df, self.database_df)
            if len(list(distance_dfs.values())[0]) > 0:
                self.generate_summary(distance_dfs, f'{self.sheet_name}_dist')

            if self.best_setup_for_session:
                best_setup_dfs = self.get_best_setup_each_session()
                if len(list(best_setup_dfs.values())[0]):
                    self.generate_summary(best_setup_dfs, f'{self.sheet_name}_1_setup_per_session')

            for feature in self.features:
                for value in self.database_df[feature].unique():
                    dfs_value = {key: df.loc[self.database_df[feature] == value]
                                 for key, df in self.data_frames.items()}
                    self.generate_summary(dfs_value, f'{self.sheet_name}_{value}')
        if (self.only_high_qual and not self.only_low_qual) or self.is_spot:
            if 'bland_altman' in self.outputs:
                for ver, pred in self.all_pred.items():
                    spot_str = 'spot' if self.is_spot else 'dyanmic'
                    plot_bland_altman(
                        self.all_gt[ver], pred, self.vs,
                        self.full_file_path.replace('.xlsx', f'_{spot_str}_{self.vs}_{ver}_bland_altman.png'),
                        title='', x_axis_ref=True)
                self.all_pred = dict()
                self.all_gt = dict()

        if ('pptx' in self.outputs and not self.only_high_qual and not self.only_low_qual
                and 'spot' not in self.sheet_name):
            self.run_presentations()

    def evaluate(self):
        for args in self.args_list:
            self.update_args(args.__dict__)
            first_folder = self.folder_list[0]
            is_spot = 'spot' in [os.path.basename(first_folder), os.path.dirname(first_folder)]
            self.is_spot = is_spot
            self.run_single_vs_single_settings()

        self.summary_hiqh_qual_low_qual_spot_all_df = add_tilda(self.summary_hiqh_qual_low_qual_spot_all_df)
        self.save_sum_to_excel(self.summary_hiqh_qual_low_qual_spot_all_df, f'summary_all_{self.vs}', sheet_index=1)

    def load_ref(self):
        if self.merge_benchmarks:
            setups = get_setups_union_folders(self.folder_list, self.vs)
        else:
            setups = get_setups_intersect_folders(self.folder_list, self.vs)
        if self.setups is not None:
            setups = intersect((self.setups, setups))
        if not self.force:
            setups = intersect((setups, get_not_invalid_setups(self.db)))

        assert len(setups), f'no setups found in {join(self.folder_list)}'
        self.ref_holder.load_ref(setups, self.vs)
        self.relevant_setups = intersect((setups, self.ref_holder.get_setups_list(self.vs)))

    @staticmethod
    def loading_folder(fol):
        base, mode = os.path.split(remove_sep_at_end(fol))
        return os.path.join(base, mode.replace('spot', 'dynamic'))

    def load_pred(self):
        for ver, folder in zip(self.version_names, self.folder_list):
            if not (ver in self.pred and self.merge_benchmarks):
                self.pred[ver] = dict()
                self.high_qual[ver] = dict()
            for num in intersect([self.relevant_setups, get_list_of_setups_in_folder_from_vs(folder, self.vs)]):
                loaded = load_pred_high_qual(folder, num, self.vs)
                self.pred[ver][num] = loaded['pred']
                self.high_qual[ver][num] = loaded.get('high_quality')


class RREvaluator(HREvaluator):
    """ RR eval is just like HR"""


class IEEvaluator(HREvaluator):
    def __init__(self, *args):
        super().__init__(*args)
        self.sheet_name = self.vs

    @staticmethod
    def loading_folder(fol):
        return fol

    def match_lists(self, v_name, setup):
        pred = self.pred[v_name][setup]
        gt = self.ref_holder.get_reference(self.vs, setup)
        if self.match_list_type == 'no':
            min_len = min(len(gt), len(pred))
            pred, gt = pred[:min_len], gt[:min_len]
            shift = 0
        elif self.match_list_type == 'mse':
            pred, gt, shift = match_lists(pred, gt)
        else:
            pred, gt, shift = match_lists_by_ts(pred, gt, setup_id=setup, vs=self.vs, db=db)
            min_len = min(len(gt), len(pred))
            pred, gt = pred[:min_len], gt[:min_len]
        if len(pred) == 0:
            raise ValueError(f'matched pred and ref is empty. setup {setup}, vital sign: {self.vs}')
        return pd.DataFrame({'pred': pred, 'gt': gt}), shift

    def run_single_vs_single_settings(self):
        if 'xlsx' in self.outputs:
            self.run_excels()

    def evaluate(self):
        self.run_single_vs_single_settings()

    def plot_and_compare_selected_setups(self, dfs):
        return


class STATEvaluator(IEEvaluator):
    def __init__(self, *args):
        super().__init__(*args)
        self.sheet_name = f'stat_trans_{self.transition_length}'

    def get_valid_indices(self, pred_ref):
        valid_indices = np.array(pred_ref['pred']) != -1
        if self.transition_length > 0:
            for i, diff in enumerate(np.diff(pred_ref['gt'])):
                if diff != 0:
                    valid_indices[i: i + self.transition_length] = False
        return valid_indices

    @staticmethod
    def confusion_matrix_eval(pred_ref):
        results = pd.Series()
        conf_mat = confusion_matrix(
            pred_ref['gt'], pred_ref['pred'], labels=np.arange(len(STAT_CLASSES)))
        conf_mat_flat = conf_mat.flatten()
        i = 0
        for ref in STAT_CLASSES:
            for pred in STAT_CLASSES:
                results[f'Number of points ref={ref}, pred={pred}'] = conf_mat_flat[i]
                i += 1
        return results

    @staticmethod
    def events_eval(pred_ref):
        results = pd.Series()
        events = list(np.where(list(np.diff(pred_ref['gt']) != 0))[0]) + [len(pred_ref['gt'])]
        for i, event in enumerate(events[:-1]):
            event_length = min(events[i + 1] - event, 60)
            event_ref = pred_ref['gt'][event + 1]
            if event_length < 20 or (event_ref == 1 and pred_ref['gt'][event] == 2):
                continue
            pred_during_event = pred_ref['pred'][event: event + event_length]
            indices_detected = np.where(pred_during_event == event_ref)[0]
            detected_points = len(indices_detected)
            detection_time = indices_detected[0] if detected_points else np.nan
            base_col_name = STAT_CLASSES[event_ref]
            results[f'{base_col_name} detection time'] = detection_time
            results[f'{base_col_name} detection points'] = detected_points
            results[f'{base_col_name} total event duration'] = event_length
        return results

    def evaluate_single_setup(self, setup, ver):
        pred_gt, shift = self.match_lists(ver, setup)
        valid_indices = self.get_valid_indices(pred_gt)
        pred_gt_no_transition = pred_gt[valid_indices]
        setup_results = pd.Series({'Setup #': setup,
                                   'number of time points': len(pred_gt_no_transition)})
        return pd.concat((setup_results, self.confusion_matrix_eval(pred_gt_no_transition), self.events_eval(pred_gt)))

    def summarize_version_category(self, df_with_db_cols: pd.DataFrame):
        n_states = len(STAT_CLASSES)
        columns_to_remove = set(df_with_db_cols.columns) & (
                set(self.database_df.columns) | {'Setup #', 'number of time points', 'Plot'})
        df = df_with_db_cols.drop(columns_to_remove, axis=1)
        summary = df.sum()
        metrics_summary = pd.Series()

        for class_stat, label in enumerate(STAT_CLASSES):
            if f'{label} detection points' in df:
                metrics_summary[f'{label} Detection Rate'] = np.mean(df[f'{label} detection points'].dropna() > 0)
                metrics_summary[f'{label} Mean Detection Time'] = np.nanmean(df[f'{label} detection time'])
                summary.drop([f'{label} detection points', f'{label} detection time'], inplace=True)
            metrics_summary[f'Percentage of setups with false {label}'] = np.mean(np.logical_or(
                *[(df.iloc[:, class_stat + n_states * i] > 0).values for i in set(range(n_states)) - {class_stat}]))

        for ind_ref, stat_ref in enumerate(STAT_CLASSES):
            for ind_pred, stat_pred in enumerate(STAT_CLASSES):
                metrics_summary[f'Percent pred={stat_pred} given ref={stat_ref}'] = \
                    summary[ind_ref * n_states + ind_pred] / (1e-8 + summary[n_states * ind_ref:
                                                                             n_states * ind_ref + n_states].sum())

        return pd.concat([metrics_summary, summary])


class BBIEvaluator(IEEvaluator):
    def __init__(self, wrapper, l_args, _):
        args = l_args[-1]
        if args.setups is None:
            args.setups = wrapper.db.setup_by_target(Target.back)
        else:
            args.setups = intersect((wrapper.db.setup_by_target(Target.back), args.setups))
        super().__init__(wrapper, [args], 'bbi')

    def summarize_version_category(self, df):
        summary = pd.Series()
        df = df[df['n peaks reference'] / df['duration'] > 0.5]
        for tolerance in BBI_TOLERANCE:
            fpr = df[f'false positives tolerance {tolerance} ms'] / df['n peaks reference']
            tpr = df[f'true positives tolerance {tolerance} ms'] / df['n peaks reference']
            summary[f'fpr < 15% and tpr > 90% tolerance {tolerance} ms per setup'] = \
                np.average(np.logical_and(fpr < 0.15, tpr > 0.9))
        return summary

    def evaluate_single_setup(self, setup, ver):
        pred = self.pred[ver][setup]
        gt = self.ref_holder.get_reference(self.vs, setup)
        gt_ts = db.setup_ts(setup, Sensor.epm_10m)
        if len(gt) > 1 and len(pred) > 1 and gt_ts is not None:
            hb_ref = BBIP.Heartbeats(gt, gt_ts['start'])
            hb_est = BBIP.Heartbeats(pred, db.setup_ts(setup, Sensor.nes)['start'])
            try:
                setup_results = BBIP.performance(hb_ref, hb_est, BBI_TOLERANCE)
            except IndexError:  # can't crop
                return
            return pd.concat([pd.Series({'Setup #': setup}), pd.Series(setup_results)])


def close_xlsx(writer_obj, workbook):
    excel_path = writer_obj.handles.handle.name
    if len(workbook.sheetnames) > 1:
        writer_obj.save()
        os.rename(excel_path, excel_path.replace('_in_progress.xlsx', '.xlsx'))
    else:
        os.remove(excel_path)


if __name__ == '__main__':

    matplotlib.use('Agg')
    db = DB()
    original_args = get_args()
    assert np.all([os.path.exists(f) for f in original_args.folder_list]) or (len(original_args.folder_list) == 1
                                                                              and original_args.setups)
    collect_result(original_args.setups, original_args.folder_list[0])
    list_of_args = get_list_of_args(original_args)
    if 'xlsx' in original_args.outputs:
        writer, xl_workbook = open_excel_and_get_writer(list_of_args[0].full_file_path, original_args.dont_overwrite)
    reference_holder_object = ReferenceHolder(db, original_args.force or original_args.force_ref)
    evaluator = EvaluatorWrapper(db, reference_holder_object)
    try:
        for vital_sign in list_of_args[0].vital_signs:
            evaluator.eval_single_vs(list_of_args, vital_sign)
    except BaseException as e:  # noqa
        print('excel saved in the middle of the run')
        traceback.print_exc()
    finally:
        if 'xlsx' in original_args.outputs:
            close_xlsx(writer, xl_workbook)
